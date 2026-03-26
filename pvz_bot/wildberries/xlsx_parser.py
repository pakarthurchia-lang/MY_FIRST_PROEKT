"""
Парсер XLSX-отчёта Wildberries ПВЗ.

Wildberries предоставляет отчёты о начислениях для операторов ПВЗ
через личный кабинет pvz.wildberries.ru → Финансы → Отчёты.

Типичная структура отчёта WB ПВЗ:
  Столбцы: Название ПВЗ | Вознаграждение | Удержания/Штрафы | Итого
  Может содержать несколько листов: начисления, детализация, штрафы.

Как скачать отчёт:
  1. Зайди на pvz.wildberries.ru
  2. Финансы → Отчёт о начислениях
  3. Выбери период → Скачать (XLSX)
"""

import io
from typing import Dict, Optional


def _find_col(row_lower: list, *keywords) -> Optional[int]:
    """Ищет индекс первого столбца, содержащего все ключевые слова."""
    for i, cell in enumerate(row_lower):
        if all(kw in cell for kw in keywords):
            return i
    return None


def parse_wb_xlsx(file_bytes: bytes) -> Dict[str, dict]:
    """
    Парсит XLSX-отчёт начислений Wildberries ПВЗ.

    Возвращает:
        {pvz_name: {"revenue": float, "fines": float, "orders": int}}

    revenue — вознаграждение (до вычета штрафов)
    fines   — удержания/штрафы (положительное число)
    orders  — количество выдач (0 если не найдено)
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Ищем лист с основными начислениями
    ws = None
    priority_keywords = ["начислен", "выплат", "финанс", "отчёт", "report", "accrual"]
    for name in wb.sheetnames:
        nl = name.lower()
        if any(kw in nl for kw in priority_keywords):
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # Ищем строку заголовков (первые 15 строк)
    col_pvz_name = None
    col_revenue = None
    col_fines = None
    col_orders = None
    header_row_idx = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        if not row:
            continue
        row_lower = [str(c).lower().strip() if c is not None else "" for c in row]

        # Название ПВЗ
        pvz_col = (
            _find_col(row_lower, "назван") or
            _find_col(row_lower, "наименован") or
            _find_col(row_lower, "адрес") or
            _find_col(row_lower, "пвз") or
            _find_col(row_lower, "pvz")
        )

        # Вознаграждение / выручка
        rev_col = (
            _find_col(row_lower, "вознагражден") or
            _find_col(row_lower, "начислен") or
            _find_col(row_lower, "итого") or
            _find_col(row_lower, "сумм") or
            _find_col(row_lower, "выплат")
        )

        # Штрафы / удержания
        fine_col = (
            _find_col(row_lower, "штраф") or
            _find_col(row_lower, "удержан") or
            _find_col(row_lower, "вычет") or
            _find_col(row_lower, "санкц")
        )

        # Количество выдач
        ord_col = (
            _find_col(row_lower, "выдач") or
            _find_col(row_lower, "количест") or
            _find_col(row_lower, "кол") or
            _find_col(row_lower, "посылок") or
            _find_col(row_lower, "заказ")
        )

        if pvz_col is not None and rev_col is not None:
            col_pvz_name = pvz_col
            col_revenue = rev_col
            col_fines = fine_col
            col_orders = ord_col
            header_row_idx = i
            break

    # Фолбэк: стандартный порядок колонок WB ПВЗ отчёта
    if col_pvz_name is None:
        col_pvz_name = 1
    if col_revenue is None:
        col_revenue = 3

    result: Dict[str, dict] = {}
    data_start = (header_row_idx + 1) if header_row_idx else 2

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row:
            continue
        max_needed = max(col_pvz_name, col_revenue,
                         col_fines if col_fines else 0,
                         col_orders if col_orders else 0)
        if len(row) <= max_needed:
            continue

        pvz_name = row[col_pvz_name]
        if not pvz_name:
            continue
        pvz_name = str(pvz_name).strip()
        if not pvz_name or pvz_name.lower() in ("итого", "total", "всего"):
            continue

        try:
            revenue = float(row[col_revenue]) if row[col_revenue] is not None else 0.0
        except (ValueError, TypeError):
            continue

        fines = 0.0
        if col_fines is not None and col_fines < len(row) and row[col_fines] is not None:
            try:
                raw_fine = float(row[col_fines])
                fines = abs(raw_fine)  # WB может хранить как отрицательное
            except (ValueError, TypeError):
                pass

        orders = 0
        if col_orders is not None and col_orders < len(row) and row[col_orders] is not None:
            try:
                orders = int(float(row[col_orders]))
            except (ValueError, TypeError):
                pass

        if pvz_name in result:
            result[pvz_name]["revenue"] += revenue
            result[pvz_name]["fines"] += fines
            result[pvz_name]["orders"] += orders
        else:
            result[pvz_name] = {"revenue": revenue, "fines": fines, "orders": orders}

    return result


def extract_period_from_wb_xlsx(file_bytes: bytes) -> Optional[dict]:
    """
    Пытается извлечь период (месяц/год) из названия листа или первых строк файла.
    Возвращает {"month": int, "year": int} или None.
    """
    import openpyxl
    import re

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        # Ищем дату в именах листов
        MONTHS_MAP = {
            "январ": 1, "феврал": 2, "март": 3, "апрел": 4,
            "май": 5, "июн": 6, "июл": 7, "август": 8,
            "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
        }
        for sheet_name in wb.sheetnames:
            sn = sheet_name.lower()
            for key, month_num in MONTHS_MAP.items():
                if key in sn:
                    # Ищем год рядом
                    years = re.findall(r"20\d{2}", sn)
                    year = int(years[0]) if years else None
                    if year:
                        return {"month": month_num, "year": year}

        # Ищем дату в первых строках активного листа
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                cell_str = str(cell)
                years = re.findall(r"20\d{2}", cell_str)
                if years:
                    for key, month_num in MONTHS_MAP.items():
                        if key in cell_str.lower():
                            return {"month": month_num, "year": int(years[0])}
    except Exception:
        pass
    return None
