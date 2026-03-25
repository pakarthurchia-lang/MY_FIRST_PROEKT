"""
Парсер XLSX-отчёта Яндекс Маркет (раздел month-reports, вкладка MONTH_CLOSING_BILLING).

Структура файла:
  Столбцы: ID ПВЗ | Название Г | Услуга | Сумма
  Строки: по одной позиции на каждую услугу/ПВЗ
  Сумма может быть положительной (выплата) или отрицательной (удержание).

Раздел транзакций (отдельный лист):
  Столбцы: ID ПВЗ | Название Г | Услуга | Время | Заказ/Воз | Товар |
           Стоимость заказа за 1шт | ... | Количество | Тип оплаты | Стоимость | Тариф | ...
  Стоимость — тариф за каждую посылку (оборот ПВЗ).
"""

import io
from typing import Dict, Optional


def _find_col(row_lower: list, *keywords) -> Optional[int]:
    """Ищет индекс первого столбца, содержащего все ключевые слова."""
    for i, cell in enumerate(row_lower):
        if all(kw in cell for kw in keywords):
            return i
    return None


def parse_ym_turnover(file_bytes: bytes) -> Dict[str, float]:
    """
    Парсит лист транзакций XLSX Яндекс Маркет.
    Возвращает {pvz_name: сумма_по_колонке_Стоимость} — оборот по каждому ПВЗ.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Ищем лист транзакций
    ws = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "транзакц" in nl or "transaction" in nl or "детал" in nl:
            ws = wb[name]
            break
    # Если не нашли по имени — ищем лист с колонками "название" + "стоимость"
    if ws is None:
        for name in wb.sheetnames:
            sheet = wb[name]
            for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                row_lower = [str(c).lower().strip() if c else "" for c in row]
                if _find_col(row_lower, "назван") is not None and _find_col(row_lower, "стоимост") is not None:
                    ws = sheet
                    break
            if ws is not None:
                break
    if ws is None:
        ws = wb.active

    # Находим заголовки
    col_pvz_name = None
    col_cost = None
    header_row_idx = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if row is None:
            continue
        row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
        pvz_col = _find_col(row_lower, "назван")
        # Ищем "Стоимость" — но НЕ "стоимость заказа" (это цена товара, не тариф)
        cost_col = None
        for j, cell in enumerate(row_lower):
            if cell == "стоимость":  # точное совпадение — тариф за посылку
                cost_col = j
                break
        if cost_col is None:
            for j, cell in enumerate(row_lower):
                if "стоимост" in cell and "заказ" not in cell and "товар" not in cell and "ндс" not in cell:
                    cost_col = j
                    break
        if pvz_col is not None and cost_col is not None:
            col_pvz_name = pvz_col
            col_cost = cost_col
            header_row_idx = i
            break

    if col_pvz_name is None:
        col_pvz_name = 1  # фолбэк: Название Г
    if col_cost is None:
        col_cost = 10     # фолбэк: 11-я колонка

    result: Dict[str, float] = {}
    data_start = (header_row_idx + 1) if header_row_idx else 2

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if row is None or len(row) <= max(col_pvz_name, col_cost):
            continue
        pvz_name = row[col_pvz_name]
        cost = row[col_cost]
        if not pvz_name:
            continue
        pvz_name = str(pvz_name).strip()
        try:
            cost = float(cost) if cost is not None else 0.0
        except (ValueError, TypeError):
            continue
        result[pvz_name] = result.get(pvz_name, 0.0) + cost

    return result


def parse_ym_fines(file_bytes: bytes, month: int = None, year: int = None) -> Dict[str, dict]:
    """
    Парсит лист «Баллы подлежащие вычету» XLSX Яндекс Маркет.
    1 балл = 1 рубль.

    Возвращает {pvz_name: {"total": float, "items": [{"date", "reason", "amount"}]}}
    Если month/year заданы — фильтрует по дате применения.
    """
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Ищем лист с баллами/штрафами
    ws = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "балл" in nl or "вычет" in nl or "штраф" in nl or "penalty" in nl or "fine" in nl:
            ws = wb[name]
            break
    if ws is None:
        return {}

    # Находим заголовки
    col_pvz_name = None
    col_date = None
    col_reason = None
    col_amount = None
    header_row_idx = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if not row:
            continue
        row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
        for j, cell in enumerate(row_lower):
            if "назван" in cell and col_pvz_name is None:
                col_pvz_name = j
            if (cell.startswith("дата") or cell == "date") and col_date is None:
                col_date = j
            if ("качеств" in cell or "причин" in cell or "reason" in cell) and col_reason is None and j != col_date:
                col_reason = j
            if ("сумм" in cell or "балл" in cell or "amount" in cell) and col_amount is None:
                col_amount = j
        if col_pvz_name is not None and col_amount is not None:
            header_row_idx = i
            break

    if col_pvz_name is None:
        col_pvz_name = 1
    if col_date is None:
        col_date = 2
    if col_reason is None:
        col_reason = 3
    if col_amount is None:
        col_amount = 4

    result: Dict[str, dict] = {}
    data_start = (header_row_idx + 1) if header_row_idx else 2

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or len(row) <= max(col_pvz_name, col_amount):
            continue

        pvz_name = row[col_pvz_name]
        if not pvz_name:
            continue
        pvz_name = str(pvz_name).strip()

        # Фильтрация по месяцу/году
        if month and year and col_date is not None and col_date < len(row):
            raw_date = row[col_date]
            try:
                if isinstance(raw_date, str):
                    dt = datetime.strptime(raw_date[:10], "%Y-%m-%d")
                elif hasattr(raw_date, "month"):
                    dt = raw_date
                else:
                    dt = None
                if dt and (dt.month != month or dt.year != year):
                    continue
            except (ValueError, TypeError):
                pass

        try:
            amount = float(row[col_amount]) if row[col_amount] is not None else 0.0
        except (ValueError, TypeError):
            continue

        if amount == 0:
            continue

        reason = ""
        if col_reason is not None and col_reason < len(row) and row[col_reason]:
            reason = str(row[col_reason]).strip()

        date_str = ""
        if col_date is not None and col_date < len(row) and row[col_date]:
            raw = row[col_date]
            try:
                if isinstance(raw, str):
                    date_str = raw[:10]
                elif hasattr(raw, "strftime"):
                    date_str = raw.strftime("%Y-%m-%d")
            except Exception:
                pass

        if pvz_name not in result:
            result[pvz_name] = {"total": 0.0, "items": []}
        result[pvz_name]["total"] += amount
        result[pvz_name]["items"].append({"date": date_str, "reason": reason, "amount": amount})

    return result


def parse_ym_xlsx(file_bytes: bytes) -> Dict[str, float]:
    """
    Принимает байты XLSX-файла.
    Возвращает словарь {pvz_name: total_amount} —
    суммарное вознаграждение по каждому ПВЗ.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Пробуем найти нужный лист по названию, иначе берём первый
    ws = None
    for name in wb.sheetnames:
        if "billing" in name.lower() or "закрыт" in name.lower() or "month" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # Найдём заголовки (первая строка с данными)
    header_row = None
    col_pvz_name = None
    col_amount = None

    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row is None:
            continue
        row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
        # Ищем столбец с названием ПВЗ
        for i, cell in enumerate(row_lower):
            if "назван" in cell or "название" in cell:
                col_pvz_name = i
            if "сумм" in cell or "amount" in cell:
                col_amount = i
        if col_pvz_name is not None and col_amount is not None:
            header_row = row
            break

    # Если заголовки не найдены — фолбэк: стандартный порядок столбцов
    # ID ПВЗ(0) | Название Г(1) | Услуга(2) | Сумма(3)
    if col_pvz_name is None:
        col_pvz_name = 1
    if col_amount is None:
        col_amount = 3

    result: Dict[str, float] = {}

    # Если заголовок не найден, начинаем с первой строки
    data_start_row = 2 if header_row is not None else 1

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if row is None:
            continue
        # Строка слишком короткая
        if len(row) <= max(col_pvz_name, col_amount):
            continue

        pvz_name = row[col_pvz_name]
        amount = row[col_amount]

        if pvz_name is None or pvz_name == "":
            continue
        pvz_name = str(pvz_name).strip()
        if not pvz_name:
            continue

        try:
            amount = float(amount) if amount is not None else 0.0
        except (ValueError, TypeError):
            continue

        result[pvz_name] = result.get(pvz_name, 0.0) + amount

    return result
